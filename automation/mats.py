#!/usr/bin/env python
# coding: utf-8

from docx import Document
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from openpyxl import load_workbook
import os
import re

def make_docx(template, newfilename, context):
    doc = DocxTemplate(template)
    doc.render(context)
    doc.save(newfilename)

def parse_excel_data(in_file):
    """Parses excel spreadsheet.
    returns a list of dictionaries with column headers as keys
    """

    # Load data from active workbook sheet
    open_sheet = load_workbook(in_file).active
    columns = tuple(open_sheet.iter_rows(min_row=0, max_row=1, values_only=True))[0]

    dictlist = [
        dict(zip(columns, row))
        for row in open_sheet.iter_rows(min_row=2, values_only=True)
    ]
    return dictlist

def combine_all_docx(filename_master, files_list):
    from datetime import datetime

    number_of_sections = len(files_list)
    master = Document(filename_master)
    composer = Composer(master)
    for i in range(0, number_of_sections):
        doc_temp = Document(files_list[i])
        composer.append(doc_temp)
    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    outputfile = os.path.join(r"..", f"mats_{now}.docx")
    composer.save(outputfile)
    print(outputfile)

context = {
    "medication": "",
    "ref": "",
    "class_t": "",
    "class_P": "",
    "safe_dose": "",
    "action": "",
    "indication": "",
    "assessments": "",
    "contraindictations": "",
    "side_effects": "",
    "education": "",
    "rate_of_admin": "",
    "dilution": "",
}

outputs = []
for med in parse_excel_data(os.path.join("..", "mats.xlsx")):
    for col in med:
        if med[col] is not None:
            context[col] = med[col]
        else:
            context[col] = ""
    med_sanitized = re.sub('[^0-9a-zA-Z]+', '', med["medication"])
    output = os.path.join(r"..", "MAT" + str(med_sanitized) + ".docx")

    make_docx("mat_template.docx", output, context)
    outputs.append(output)

# insert page break after every 2 mats
for i in range(len(outputs) + (int(len(outputs) / 2))):
    if (i + 1) % 3 == 0:
        outputs.insert(i, "pagebreak.docx")

# print(outputs)
combine_all_docx("merge.docx", outputs)

# clean up intermediate files
for doc in outputs:
    if doc != "pagebreak.docx":
        os.remove(doc)
